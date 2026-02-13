import pytest
import openpyxl
from io import BytesIO
from datetime import date
from django.contrib.auth import get_user_model
from apps.core.utils.excel_generator import create_excel_sample_file
from apps.core.utils.sample_headers import INVESTOR_HEADERS, INVESTOR_CHOICES, SCHEME_HEADERS, SCHEME_CHOICES
from apps.users.utils.parsers import import_investors_from_file
from apps.users.models import InvestorProfile
from apps.products.utils.parsers import import_schemes_from_file
from apps.products.models import Scheme

User = get_user_model()

@pytest.mark.django_db
def test_investor_import_roundtrip():
    # 1. Generate Sample File
    excel_file = create_excel_sample_file(INVESTOR_HEADERS, INVESTOR_CHOICES)

    # 2. Add Data using openpyxl
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Fill Row 2
    row = 2
    data_map = {
        'PAN': 'ABCDE1234F',
        'First Name': 'John',
        'Last Name': 'Doe',
        'Email': 'john@example.com',
        'Mobile': '9876543210',
        'Tax Status': 'Individual', # Should map to '01'
        'Occupation': 'Service',    # Should map to '02'
        'Gender': 'Male',
        'Date of Birth (YYYY-MM-DD)': '1990-01-01',
        'Is Offline (Y/N)': 'Yes',
        'Bank 1 Account No': '1234567890',
        'Bank 1 IFSC': 'HDFC0001234',
        'Nominee 1 Name': 'Jane Doe',
        'Nominee 1 %': 100
    }

    # Helper to find col index
    headers = [cell.value for cell in ws[1]]

    for header, value in data_map.items():
        if header in headers:
            col_idx = headers.index(header) + 1
            ws.cell(row=row, column=col_idx, value=value)

    # Save back to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    output.name = 'test_investor.xlsx' # Important for parser detection

    # 3. Run Import
    count, errors = import_investors_from_file(output)

    assert not errors, f"Import failed with errors: {errors}"
    assert count == 1

    # 4. Verify DB
    investor = InvestorProfile.objects.get(pan='ABCDE1234F')
    assert investor.firstname == 'John'
    assert investor.lastname == 'Doe'
    assert investor.email == 'john@example.com'
    assert investor.tax_status == '01' # Individual
    assert investor.occupation == '02' # Service
    assert investor.gender == 'M'
    assert investor.dob == date(1990, 1, 1)
    assert investor.is_offline is True

    # Verify User
    user = User.objects.get(username='ABCDE1234F')
    assert user.email == 'john@example.com'

    # Verify Bank
    bank = investor.bank_accounts.first()
    assert bank.account_number == '1234567890'
    assert bank.ifsc_code == 'HDFC0001234'

    # Verify Nominee
    nominee = investor.nominees.first()
    assert nominee.name == 'Jane Doe'
    assert nominee.percentage == 100.00

@pytest.mark.django_db
def test_scheme_import_roundtrip():
    # 1. Generate Sample File
    excel_file = create_excel_sample_file(SCHEME_HEADERS, SCHEME_CHOICES)

    # 2. Add Data
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    row = 2
    data_map = {
        'Scheme Code': 'SCHEME001',
        'Scheme Name': 'Test Scheme Growth',
        'AMC Code': 'HDFC_MF',
        'Scheme Type': 'Open Ended',
        'Category': 'Equity',
        'ISIN': 'INF123456789',
        'Unique No': 1001,
        'Purchase Allowed (Y/N)': 'Yes',
        'Min Purchase Amount': 5000,
        'SIP Allowed (Y/N)': 'Yes',
        'Start Date': '2023-01-01'
    }

    headers = [cell.value for cell in ws[1]]
    for header, value in data_map.items():
        if header in headers:
            col_idx = headers.index(header) + 1
            ws.cell(row=row, column=col_idx, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    output.name = 'test_scheme.xlsx'

    # 3. Run Import
    count, errors = import_schemes_from_file(output)

    assert not errors, f"Import failed: {errors}"
    assert count == 1

    # 4. Verify DB
    scheme = Scheme.objects.get(scheme_code='SCHEME001')
    assert scheme.name == 'Test Scheme Growth'
    assert scheme.amc.code == 'HDFC_MF'
    assert scheme.purchase_allowed is True
    assert scheme.min_purchase_amount == 5000.00
    assert scheme.is_sip_allowed is True
    assert scheme.start_date == date(2023, 1, 1)
