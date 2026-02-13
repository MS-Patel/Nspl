import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from io import BytesIO

def create_excel_sample_file(headers, choices=None):
    """
    Generates an Excel file with headers and data validation dropdowns.

    :param headers: List of column headers (strings).
    :param choices: Dictionary where keys are header names and values are lists of choices.
    :return: BytesIO object containing the Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"

    # Create a hidden sheet for validation lists if choices exist
    validation_sheet = None
    if choices:
        validation_sheet = wb.create_sheet("ValidationLists")
        validation_sheet.sheet_state = 'hidden'

    # 1. Write Headers and Setup Columns
    header_col_map = {}
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        # Adjust column width
        ws.column_dimensions[get_column_letter(col_idx)].width = len(header) + 5
        header_col_map[header] = col_idx

    # 2. Add Data Validation
    if choices and validation_sheet:
        val_col_idx = 1

        for header, options in choices.items():
            if header not in header_col_map or not options:
                continue

            col_idx = header_col_map[header]
            col_letter = get_column_letter(col_idx)

            # Write options to the hidden sheet
            val_col_letter = get_column_letter(val_col_idx)

            # Write header for clarity (optional, row 1)
            validation_sheet.cell(row=1, column=val_col_idx, value=header)

            for i, option in enumerate(options, start=2):
                validation_sheet.cell(row=i, column=val_col_idx, value=str(option))

            # Define the range for validation
            # E.g., 'ValidationLists'!$A$2:$A$10
            end_row = len(options) + 1
            formula = f"'ValidationLists'!${val_col_letter}$2:${val_col_letter}${end_row}"

            # Create DataValidation object
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)

            # Apply to the column (rows 2 to 1000 for sample)
            dv.add(f"{col_letter}2:{col_letter}1000")

            ws.add_data_validation(dv)

            val_col_idx += 1

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
